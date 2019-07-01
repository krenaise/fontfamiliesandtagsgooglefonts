import json
from openpyxl import Workbook

with open('families.json', 'r+') as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active

cur_row = 1

for font in data:
    for i in range(max(len(font['tags']), 5)):
        family = font['family']
        tags = font['tags']
        ws.append([])
        if i == 0:
            ws.cell(row=cur_row, column=1).value = family
        if len(tags) > i:
            ws.cell(row=cur_row, column=2).value = tags[i]
        cur_row += 1

wb.save(filename='result.xlsx')